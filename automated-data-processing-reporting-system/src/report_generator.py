"""
Generates a polished Excel report (with charts) and PNG chart summaries
from the data stored in MySQL.

Run directly:
    python -m src.report_generator
"""

import os
from datetime import datetime

import matplotlib
matplotlib.use("Agg")  # headless rendering, no display needed
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

from src.config_loader import load_config
from src.db_connector import MySQLConnector
from src.logger import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)


def _write_df_to_sheet(ws, df, start_row=1, title=None):
    """Write a DataFrame to a worksheet with basic styling, returns end row."""
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = TITLE_FONT
        row += 2

    header_row = row
    for c_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_idx, record in enumerate(df.itertuples(index=False), start=header_row + 1):
        for c_idx, value in enumerate(record, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for c_idx, col_name in enumerate(df.columns, start=1):
        width = max(12, len(col_name) + 2)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    return header_row, header_row + len(df)


def build_excel_report(db: MySQLConnector, output_path: str, top_n: int = 10) -> str:
    """Query MySQL, build a multi-sheet Excel workbook with embedded charts."""
    region_df = db.run_query("""
        SELECT region,
               SUM(total_orders) AS total_orders,
               SUM(total_quantity) AS total_quantity,
               SUM(total_revenue) AS total_revenue
        FROM monthly_region_summary
        GROUP BY region
        ORDER BY total_revenue DESC
    """)

    category_df = db.run_query("""
        SELECT product_category,
               COUNT(*) AS total_orders,
               SUM(quantity) AS units_sold,
               SUM(total_amount) AS total_revenue
        FROM sales_transactions
        GROUP BY product_category
        ORDER BY total_revenue DESC
    """)

    top_products_df = db.run_query("""
        SELECT product_name,
               SUM(quantity) AS units_sold,
               SUM(total_amount) AS total_revenue
        FROM sales_transactions
        GROUP BY product_name
        ORDER BY total_revenue DESC
        LIMIT :top_n
    """, params={"top_n": top_n})

    trend_df = db.run_query("""
        SELECT order_year, order_month, SUM(total_revenue) AS total_revenue
        FROM monthly_region_summary
        GROUP BY order_year, order_month
        ORDER BY order_year, order_month
    """)
    trend_df["period"] = trend_df["order_year"].astype(str) + "-" + trend_df["order_month"].astype(str).str.zfill(2)
    trend_df = trend_df[["period", "total_revenue"]]

    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Automated Sales Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    total_revenue = region_df["total_revenue"].sum() if not region_df.empty else 0
    total_orders = region_df["total_orders"].sum() if not region_df.empty else 0
    ws_summary["A4"] = "Total Revenue:"
    ws_summary["B4"] = float(total_revenue)
    ws_summary["A5"] = "Total Orders:"
    ws_summary["B5"] = int(total_orders)

    # --- Region sheet + bar chart ---
    ws_region = wb.create_sheet("By Region")
    h_row, end_row = _write_df_to_sheet(ws_region, region_df, title="Revenue by Region")
    chart = BarChart()
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Revenue"
    data = Reference(ws_region, min_col=4, min_row=h_row, max_row=end_row)
    cats = Reference(ws_region, min_col=1, min_row=h_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_region.add_chart(chart, f"F{h_row}")

    # --- Category sheet + bar chart ---
    ws_cat = wb.create_sheet("By Category")
    h_row, end_row = _write_df_to_sheet(ws_cat, category_df, title="Revenue by Product Category")
    chart2 = BarChart()
    chart2.title = "Revenue by Category"
    data2 = Reference(ws_cat, min_col=4, min_row=h_row, max_row=end_row)
    cats2 = Reference(ws_cat, min_col=1, min_row=h_row + 1, max_row=end_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws_cat.add_chart(chart2, f"F{h_row}")

    # --- Top products sheet ---
    ws_top = wb.create_sheet("Top Products")
    _write_df_to_sheet(ws_top, top_products_df, title=f"Top {top_n} Products by Revenue")

    # --- Monthly trend sheet + line chart ---
    ws_trend = wb.create_sheet("Monthly Trend")
    h_row, end_row = _write_df_to_sheet(ws_trend, trend_df, title="Monthly Revenue Trend")
    chart3 = LineChart()
    chart3.title = "Monthly Revenue Trend"
    data3 = Reference(ws_trend, min_col=2, min_row=h_row, max_row=end_row)
    cats3 = Reference(ws_trend, min_col=1, min_row=h_row + 1, max_row=end_row)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    ws_trend.add_chart(chart3, f"D{h_row}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel report saved to {output_path}")
    return output_path


def build_matplotlib_charts(db: MySQLConnector, output_dir: str) -> list:
    """Generate standalone PNG charts (useful for embedding in emails/slides)."""
    os.makedirs(output_dir, exist_ok=True)
    saved = []

    region_df = db.run_query("""
        SELECT region, SUM(total_revenue) AS total_revenue
        FROM monthly_region_summary GROUP BY region ORDER BY total_revenue DESC
    """)
    if not region_df.empty:
        plt.figure(figsize=(8, 5))
        plt.bar(region_df["region"], region_df["total_revenue"], color="#1F4E78")
        plt.title("Revenue by Region")
        plt.ylabel("Revenue")
        plt.xticks(rotation=30)
        plt.tight_layout()
        path = os.path.join(output_dir, "revenue_by_region.png")
        plt.savefig(path, dpi=150)
        plt.close()
        saved.append(path)

    trend_df = db.run_query("""
        SELECT order_year, order_month, SUM(total_revenue) AS total_revenue
        FROM monthly_region_summary GROUP BY order_year, order_month
        ORDER BY order_year, order_month
    """)
    if not trend_df.empty:
        trend_df["period"] = trend_df["order_year"].astype(str) + "-" + trend_df["order_month"].astype(str).str.zfill(2)
        plt.figure(figsize=(10, 5))
        plt.plot(trend_df["period"], trend_df["total_revenue"], marker="o", color="#2E86C1")
        plt.title("Monthly Revenue Trend")
        plt.ylabel("Revenue")
        plt.xticks(rotation=45)
        plt.tight_layout()
        path = os.path.join(output_dir, "monthly_revenue_trend.png")
        plt.savefig(path, dpi=150)
        plt.close()
        saved.append(path)

    return saved


def generate_full_report(config_path: str = "config/config.yaml") -> dict:
    config = load_config(config_path)
    db = MySQLConnector(config["database"])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    reports_dir = config["paths"]["reports_dir"]
    excel_path = os.path.join(reports_dir, f"sales_report_{timestamp}.xlsx")

    excel_file = build_excel_report(db, excel_path, top_n=config["report"]["top_n_products"])
    chart_files = build_matplotlib_charts(db, reports_dir)

    return {"excel_report": excel_file, "charts": chart_files}


if __name__ == "__main__":
    result = generate_full_report()
    print(result)
